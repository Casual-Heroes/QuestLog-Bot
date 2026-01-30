# cogs/xp.py - XP & Leveling System (FREE)
"""
XP cog for QuestLog.
Handles XP tracking, leveling, and Hero Tokens.

XP SOURCES:
- Messages (active)
- Media/images (active)
- Voice chat (active)
- Reactions (active)
- Slash commands (active)
- Gaming activity (passive)
- Invites (active)

TOKEN CONVERSION:
- Active: 15 tokens per 100 XP
- Passive: 5 tokens per 100 XP
"""

import time
import asyncio
import re
import discord
from discord.ext import commands, tasks
from discord import SlashCommandGroup

from config import (
    db_session_scope,
    logger,
    DefaultXPSettings,
    get_debug_guilds,
)
from models import (
    Guild, GuildMember, XPConfig, LevelRole, LevelUpConfig,
    XPExcludedChannel, XPExcludedRole, LevelRequirement, XPBoostEvent
)


class XPCog(commands.Cog):
    """XP and leveling system (FREE feature)."""

    def __init__(self, bot: commands.Bot):
        self.bot = bot
        # Per-guild invite cache: {guild_id: {invite_code: {uses, inviter, last_award}}}
        self.invite_cache = {}
        # Start background task to expire boost events
        self.check_expired_boost_events.start()

    xp = SlashCommandGroup(
        name="xp",
        description="XP and leveling commands",
        
    )

    # XP helper methods

    @staticmethod
    def get_xp_config(session, guild_id: int) -> dict:
        """Get XP config for guild, or return defaults."""
        config = session.get(XPConfig, guild_id)
        if config:
            return {
                "xp_enabled": config.xp_enabled,
                "message_xp": config.message_xp,
                "media_multiplier": config.media_multiplier,
                "reaction_xp": config.reaction_xp,
                "voice_xp": config.voice_xp_per_interval,
                "command_xp": config.command_xp,
                "gaming_xp": config.gaming_xp_per_interval,
                "invite_xp": config.invite_xp,
                "join_xp": config.join_xp,
                "tokens_active": config.tokens_per_100_xp_active,
                "tokens_passive": config.tokens_per_100_xp_passive,
                "message_cooldown": config.message_cooldown,
                "media_cooldown": config.media_cooldown,
                "reaction_cooldown": config.reaction_cooldown,
                "voice_interval": config.voice_interval,
                "gaming_interval": config.gaming_interval,
                "command_cooldown": config.command_cooldown,
                "game_launch_cooldown": config.game_launch_cooldown,
                "max_level": config.max_level,
                # XP source toggles
                "track_messages": getattr(config, "track_messages", False),
                "track_media": getattr(config, "track_media", False),
                "track_reactions": getattr(config, "track_reactions", False),
                "track_voice": getattr(config, "track_voice", False),
                "track_gaming": getattr(config, "track_gaming", False),
                "track_game_launch": getattr(config, "track_game_launch", False),
            }
        # Return defaults
        return {
            "xp_enabled": False,  # XP disabled by default
            "message_xp": DefaultXPSettings.MESSAGE_XP,
            "media_multiplier": DefaultXPSettings.MEDIA_MULTIPLIER,
            "reaction_xp": DefaultXPSettings.REACTION_XP,
            "voice_xp": DefaultXPSettings.VOICE_XP_PER_INTERVAL,
            "command_xp": DefaultXPSettings.COMMAND_XP,
            "gaming_xp": DefaultXPSettings.GAMING_XP_PER_INTERVAL,
            "invite_xp": DefaultXPSettings.INVITE_XP,
            "join_xp": DefaultXPSettings.JOIN_XP,
            "tokens_active": DefaultXPSettings.TOKENS_PER_100_XP_ACTIVE,
            "tokens_passive": DefaultXPSettings.TOKENS_PER_100_XP_PASSIVE,
            "message_cooldown": DefaultXPSettings.MESSAGE_COOLDOWN,
            "media_cooldown": DefaultXPSettings.MEDIA_COOLDOWN,
            "reaction_cooldown": DefaultXPSettings.REACTION_COOLDOWN,
            "voice_interval": DefaultXPSettings.VOICE_INTERVAL,
            "gaming_interval": DefaultXPSettings.GAMING_INTERVAL,
            "command_cooldown": DefaultXPSettings.COMMAND_COOLDOWN,
            "game_launch_cooldown": DefaultXPSettings.GAME_LAUNCH_COOLDOWN,
            "max_level": DefaultXPSettings.MAX_LEVEL,
            # XP source toggles (OPT-IN: disabled by default)
            "track_messages": False,
            "track_media": False,
            "track_reactions": False,
            "track_voice": False,
            "track_gaming": False,
            "track_game_launch": False,
        }

    @staticmethod
    def can_gain_xp(session, guild_id: int, member: discord.Member, channel: discord.abc.GuildChannel = None) -> bool:
        """Check if member can earn XP (not excluded by role or channel)."""
        # Check if XP system is enabled for this guild
        xp_config = session.get(XPConfig, guild_id)
        if not xp_config or not xp_config.xp_enabled:
            return False

        # Check excluded roles
        excluded_roles = (
            session.query(XPExcludedRole.role_id)
            .filter(XPExcludedRole.guild_id == guild_id)
            .all()
        )
        excluded_role_ids = {r[0] for r in excluded_roles}

        for role in member.roles:
            if role.id in excluded_role_ids:
                return False

        # Check excluded channels
        if channel:
            # Always exclude self-promo channel from XP
            from models import DiscoveryConfig
            discovery_config = session.get(DiscoveryConfig, guild_id)
            if discovery_config and discovery_config.selfpromo_channel_id == channel.id:
                return False

            excluded_channels = (
                session.query(XPExcludedChannel.channel_id)
                .filter(XPExcludedChannel.guild_id == guild_id)
                .all()
            )
            excluded_channel_ids = {c[0] for c in excluded_channels}
            if channel.id in excluded_channel_ids:
                return False

        return True

    @staticmethod
    def calculate_level(session, xp: float, max_level: int = 99) -> int:
        """Calculate level based on XP using level_requirements table."""
        try:
            level_reqs = (
                session.query(LevelRequirement)
                .order_by(LevelRequirement.level.desc())
                .all()
            )

            for req in level_reqs:
                if xp >= req.xp_required:
                    return min(req.level, max_level)

            return 0

        except Exception as e:
            logger.error(f"Error calculating level: {e}")
            # Fallback formula if table query fails
            level = 0
            while level < max_level and xp >= round(7 * ((level + 1) ** 1.5)):
                level += 1
            return level

    @staticmethod
    def calculate_xp_with_boosts(session, guild_id: int, user_id: int, base_xp: float,
                                  guild: discord.Guild = None, channel_id: int = None) -> tuple:
        """
        Calculate XP with active boost event multipliers applied.

        Args:
            session: Database session
            guild_id: Guild ID
            user_id: User ID
            base_xp: Base XP amount to award
            guild: Discord guild object (optional, for role checking)
            channel_id: Channel ID (optional, for channel-specific boosts)

        Returns:
            tuple: (final_xp, bonus_tokens, active_event_names)
        """
        from models import XPBoostEvent

        # Get all active boost events for this guild
        current_time = int(time.time())
        events = session.query(XPBoostEvent).filter(
            XPBoostEvent.guild_id == guild_id,
            XPBoostEvent.is_active == True
        ).all()

        total_multiplier = 1.0
        total_token_bonus = 0
        active_events = []

        for event in events:
            # Skip expired events (will be auto-disabled by background task)
            if event.end_time and event.end_time < current_time:
                continue

            # Check scope
            applies = False
            if event.scope == 'server':
                applies = True
            elif event.scope == 'channel' and channel_id:
                applies = (event.scope_id == channel_id)
            elif event.scope == 'role' and guild and user_id:
                # Check if user has the role
                member = guild.get_member(user_id)
                if member:
                    applies = any(role.id == event.scope_id for role in member.roles)

            if applies:
                # Stack multipliers (e.g., 2x + 1.5x = 3.5x total)
                total_multiplier += (event.multiplier - 1.0)
                total_token_bonus += (event.token_bonus or 0)
                active_events.append(event.name)

        # Always round 0.5 up (not banker's rounding) for intuitive game mechanics
        # This ensures 4.5 XP becomes 5, not 4
        final_xp = int(base_xp * total_multiplier + 0.5)
        bonus_tokens = int(final_xp * total_token_bonus / 100 + 0.5) if total_token_bonus > 0 else 0

        if active_events:
            logger.debug(
                f"XP Boost Applied: guild={guild_id}, user={user_id}, "
                f"base={base_xp}, final={final_xp}, multiplier={total_multiplier:.2f}x, "
                f"events={', '.join(active_events)}"
            )

        return final_xp, bonus_tokens, active_events

    @staticmethod
    def add_xp(session, guild_id: int, user_id: int, amount: float,
               display_name: str = None, engagement_type: str = "active",
               guild: discord.Guild = None, channel_id: int = None,
               source: str = None) -> tuple:
        """
        Add XP to a member and calculate level/token rewards.
        Now includes XP boost event support.

        Returns:
            tuple: (old_level, new_level, current_hero_tokens, token_diff)
        """
        if amount <= 0:
            return (0, 0, 0, 0)

        # Get XP config for guild
        xp_config = XPCog.get_xp_config(session, guild_id)

        source_toggle_map = {
            "messages": "track_messages",
            "media": "track_media",
            "reactions": "track_reactions",
            "voice": "track_voice",
            "gaming": "track_gaming",
            "game_launch": "track_game_launch",
        }
        if source:
            source_key = source.lower() if isinstance(source, str) else source
            toggle_key = source_toggle_map.get(source_key)
            if toggle_key and not xp_config.get(toggle_key, False):
                return (0, 0, 0, 0)

        # Apply XP boost events
        final_xp, boost_bonus_tokens, active_events = XPCog.calculate_xp_with_boosts(
            session, guild_id, user_id, amount, guild, channel_id
        )

        # Get or create member
        db_member = session.get(GuildMember, (guild_id, user_id))
        if not db_member:
            # CRITICAL: Log all new member creations to track duplicate issue
            logger.warning(
                f"CREATING NEW GUILD_MEMBER: guild_id={guild_id}, user_id={user_id}, "
                f"display_name={display_name}, user_id_type={type(user_id)}"
            )
            db_member = GuildMember(
                guild_id=guild_id,
                user_id=user_id,
                display_name=display_name,
            )
            session.add(db_member)
            session.flush()

        # Update display name if provided
        if display_name and db_member.display_name != display_name:
            db_member.display_name = display_name

        # Store old values
        old_level = db_member.level
        old_xp = db_member.xp

        # Add XP (with boost multiplier applied)
        db_member.xp += final_xp
        db_member.last_active = int(time.time())

        # Calculate hero tokens (awarded every 100 XP)
        old_hundreds = int(old_xp // 100)
        new_hundreds = int(db_member.xp // 100)
        diff = new_hundreds - old_hundreds

        token_diff = 0
        if diff > 0:
            if engagement_type == "active":
                token_diff = diff * xp_config["tokens_active"]
            else:
                token_diff = diff * xp_config["tokens_passive"]
            db_member.hero_tokens += token_diff

        # Add bonus tokens from boost events
        if boost_bonus_tokens > 0:
            db_member.hero_tokens += boost_bonus_tokens
            token_diff += boost_bonus_tokens

        # Calculate new level
        new_level = XPCog.calculate_level(session, db_member.xp, xp_config["max_level"])
        db_member.level = new_level

        logger.debug(
            f"XP Award: guild={guild_id}, user={user_id}, amount={amount}, "
            f"old_level={old_level}, new_level={new_level}, tokens={token_diff}"
        )

        return (old_level, new_level, db_member.hero_tokens, token_diff)

    async def send_level_up_notification(self, guild: discord.Guild, member: discord.Member,
                                         old_level: int, new_level: int,
                                         hero_tokens: int, token_diff: int,
                                         current_channel: discord.TextChannel = None):
        """Send level up notification respecting the guild's levelup config."""
        if new_level <= old_level:
            return

        with db_session_scope() as session:
            # Get levelup config
            levelup_config = session.get(LevelUpConfig, guild.id)
            if not levelup_config or not levelup_config.enabled:
                return

            destination = levelup_config.destination or "current"

            # Get guild for specific channel lookup
            db_guild = session.get(Guild, guild.id)
            channel_id = db_guild.level_up_channel_id if db_guild else None
            token_name = db_guild.token_name if db_guild and db_guild.token_name else "Hero Tokens"

            # Get level roles for this guild
            level_roles = (
                session.query(LevelRole)
                .filter(LevelRole.guild_id == guild.id)
                .all()
            )
            role_data = {lr.level: lr.role_id for lr in level_roles}

        # Determine destination channel based on config
        level_channel = None

        if destination == "current":
            # Send in the channel where they leveled up
            level_channel = current_channel
        elif destination == "channel":
            # Send in specific configured channel
            level_channel = guild.get_channel(channel_id) if channel_id else None
        elif destination == "dm":
            # Send as DM
            try:
                if member.dm_channel is None:
                    await member.create_dm()
                level_channel = member.dm_channel
            except Exception:
                logger.warning(f"Cannot DM user {member.id} for level-up")
                return
        elif destination == "none":
            # Silent - don't send any message
            return

        # If no valid channel found, abort
        if not level_channel:
            logger.warning(f"No valid level-up channel for guild {guild.id}, destination={destination}")
            return

        # Send level up message
        try:
            await level_channel.send(
                f"🎉 GG {member.display_name}, you just advanced to **level {new_level}**!"
            )

            if token_diff > 0:
                await level_channel.send(
                    f"🪙 You earned **{token_diff} {token_name}**! Total: **{hero_tokens}**"
                )

            # Check for level role
            if new_level in role_data:
                role = guild.get_role(role_data[new_level])
                if role and role not in member.roles:
                    try:
                        await member.add_roles(role, reason="Level milestone reached")
                        await level_channel.send(
                            f"🏆 You've earned the rank: **{role.name}**!"
                        )
                    except discord.Forbidden:
                        logger.warning(f"Cannot assign role {role.name} to {member.id}")

        except discord.Forbidden:
            logger.warning(f"Cannot send level up message for {member.id} in {level_channel}")
        except Exception as e:
            logger.error(f"Error sending level up notification: {e}", exc_info=True)

    async def check_and_award_level_roles(self, guild: discord.Guild, member: discord.Member, level: int):
        """Check and assign milestone roles based on level."""
        with db_session_scope() as session:
            level_roles = (
                session.query(LevelRole)
                .filter(LevelRole.guild_id == guild.id)
                .order_by(LevelRole.level)
                .all()
            )

            # Find appropriate role
            target_role_id = None
            remove_previous = True
            all_role_ids = []

            for lr in level_roles:
                all_role_ids.append(lr.role_id)
                if level >= lr.level:
                    target_role_id = lr.role_id
                    remove_previous = lr.remove_previous

        if not target_role_id:
            return

        target_role = guild.get_role(target_role_id)
        if not target_role:
            return

        try:
            # Remove previous milestone roles if configured
            if remove_previous:
                current_milestone_roles = [
                    role for role in member.roles
                    if role.id in all_role_ids and role.id != target_role_id
                ]
                for role in current_milestone_roles:
                    await member.remove_roles(role, reason="Upgraded to higher milestone")

            # Add new role
            if target_role not in member.roles:
                await member.add_roles(target_role, reason="Level milestone reached")

        except discord.Forbidden:
            logger.warning(f"Cannot manage roles for {member.id} in {guild.id}")
        except Exception as e:
            logger.error(f"Error managing level roles: {e}")

    # Invite tracking

    async def update_invite_cache_for_guild(self, guild: discord.Guild):
        """Update invite cache for a guild."""
        try:
            invites = await guild.invites()
            if guild.id not in self.invite_cache:
                self.invite_cache[guild.id] = {}

            for invite in invites:
                if invite.inviter and not invite.inviter.bot:
                    last_award = self.invite_cache.get(guild.id, {}).get(invite.code, {}).get("last_award", 0)
                    self.invite_cache[guild.id][invite.code] = {
                        "uses": invite.uses,
                        "inviter": invite.inviter.id,
                        "last_award": last_award,
                    }
        except discord.Forbidden:
            pass
        except Exception as e:
            logger.error(f"Error updating invite cache for {guild.name}: {e}")

    # Event listeners

    @commands.Cog.listener()
    async def on_ready(self):
        """Initialize invite cache for all guilds."""
        guilds_to_cache = []
        for guild in self.bot.guilds:
            with db_session_scope() as session:
                db_guild = session.get(Guild, guild.id)
                if db_guild and db_guild.xp_enabled:
                    guilds_to_cache.append(guild)

        # Stagger invite cache updates to avoid rate limiting
        for i, guild in enumerate(guilds_to_cache):
            if i > 0:
                await asyncio.sleep(2)  # 2 second delay between guilds
            await self.update_invite_cache_for_guild(guild)

        logger.info(f"XP Cog ready - invite cache initialized for {len(guilds_to_cache)} guilds")

    @commands.Cog.listener()
    async def on_invite_create(self, invite: discord.Invite):
        """Track new invites."""
        if not invite.guild or not invite.inviter or invite.inviter.bot:
            return

        guild_id = invite.guild.id
        if guild_id not in self.invite_cache:
            self.invite_cache[guild_id] = {}

        self.invite_cache[guild_id][invite.code] = {
            "uses": invite.uses,
            "inviter": invite.inviter.id,
            "last_award": 0,
        }

    @commands.Cog.listener()
    async def on_invite_delete(self, invite: discord.Invite):
        """Remove deleted invites from cache."""
        if not invite.guild:
            return

        guild_id = invite.guild.id
        if guild_id in self.invite_cache and invite.code in self.invite_cache[guild_id]:
            del self.invite_cache[guild_id][invite.code]

    @commands.Cog.listener()
    async def on_member_join(self, member: discord.Member):
        """Award XP for invites and joining."""
        if member.bot:
            return

        guild_id = member.guild.id

        with db_session_scope() as session:
            db_guild = session.get(Guild, guild_id)
            if not db_guild or not db_guild.xp_enabled:
                return

            xp_config = XPCog.get_xp_config(session, guild_id)
            now = int(time.time())

            # Check for invite XP
            try:
                current_invites = await member.guild.invites()
            except Exception:
                current_invites = []

            for invite in current_invites:
                cached = self.invite_cache.get(guild_id, {}).get(invite.code)
                if not cached:
                    continue

                if invite.uses > cached["uses"]:
                    # Award invite XP to inviter (once per week per inviter)
                    one_week = 7 * 24 * 60 * 60
                    if cached["last_award"] == 0 or (now - cached["last_award"]) >= one_week:
                        inviter_id = cached["inviter"]
                        inviter = member.guild.get_member(inviter_id)

                        if inviter and not inviter.bot:
                            result = XPCog.add_xp(
                                session, guild_id, inviter_id,
                                xp_config["invite_xp"],
                                inviter.display_name,
                                engagement_type="active"
                            )

                            # Update invite count
                            db_inviter = session.get(GuildMember, (guild_id, inviter_id))
                            if db_inviter:
                                db_inviter.invite_count += 1
                                db_inviter.last_invite_ts = now

                            cached["last_award"] = now
                            old_level, new_level, tokens, token_diff = result

                            # Send notification after session commits
                            if new_level > old_level:
                                session.commit()
                                await self.send_level_up_notification(
                                    member.guild, inviter,
                                    old_level, new_level, tokens, token_diff,
                                    None  # No specific channel for invite XP
                                )

                cached["uses"] = invite.uses

            # Award join XP to new member
            db_member = session.get(GuildMember, (guild_id, member.id))
            if not db_member:
                result = XPCog.add_xp(
                    session, guild_id, member.id,
                    xp_config["join_xp"],
                    member.display_name,
                    engagement_type="active"
                )
                logger.debug(f"Awarded {xp_config['join_xp']} join XP to {member.display_name}")

    @commands.Cog.listener()
    async def on_message(self, message: discord.Message):
        """Award XP for messages."""
        if message.author.bot or not message.guild:
            return

        guild_id = message.guild.id
        should_notify = False
        level_data = None

        try:
            with db_session_scope() as session:
                db_guild = session.get(Guild, guild_id)
                if not db_guild or not db_guild.xp_enabled:
                    return

                if not XPCog.can_gain_xp(session, guild_id, message.author, message.channel):
                    return

                xp_config = XPCog.get_xp_config(session, guild_id)

                # Get or create member
                db_member = session.get(GuildMember, (guild_id, message.author.id))
                if not db_member:
                    logger.warning(
                        f"CREATING NEW GUILD_MEMBER (message event): guild_id={guild_id}, "
                        f"user_id={message.author.id}, display_name={message.author.display_name}, "
                        f"user_id_type={type(message.author.id)}, username={message.author.name}"
                    )
                    db_member = GuildMember(
                        guild_id=guild_id,
                        user_id=message.author.id,
                        display_name=message.author.display_name,
                    )
                    session.add(db_member)
                    session.flush()

                if db_member.display_name != message.author.display_name:
                    db_member.display_name = message.author.display_name

                now = int(time.time())

                # Check for media
                is_media = bool(message.attachments)
                media_pattern = r"(https?://[^\s]+)|(<a?:\w+:\d+>)"
                if re.search(media_pattern, message.content):
                    is_media = True

                # Update counts - track both total messages AND media messages
                db_member.message_count += 1  # Always increment total message count
                if is_media:
                    db_member.media_count += 1  # Additionally track media messages

                # Award XP with cooldown
                if is_media:
                    # Check if media XP tracking is enabled
                    if xp_config["track_media"] and (now - db_member.last_media_ts) >= xp_config["media_cooldown"]:
                        xp_amount = xp_config["message_xp"] * xp_config["media_multiplier"]
                        result = XPCog.add_xp(
                            session, guild_id, message.author.id,
                            xp_amount, message.author.display_name, "active",
                            message.guild, message.channel.id,
                            source="media"
                        )
                        db_member.last_media_ts = now
                        old_level, new_level, tokens, token_diff = result

                        if new_level > old_level:
                            should_notify = True
                            level_data = (old_level, new_level, tokens, token_diff)
                else:
                    # Check if message XP tracking is enabled
                    if xp_config["track_messages"] and (now - db_member.last_message_ts) >= xp_config["message_cooldown"]:
                        result = XPCog.add_xp(
                            session, guild_id, message.author.id,
                            xp_config["message_xp"], message.author.display_name, "active",
                            message.guild, message.channel.id,
                            source="messages"
                        )
                        db_member.last_message_ts = now
                        old_level, new_level, tokens, token_diff = result

                        if new_level > old_level:
                            should_notify = True
                            level_data = (old_level, new_level, tokens, token_diff)

            # Send notification after commit
            if should_notify and level_data:
                await self.send_level_up_notification(
                    message.guild, message.author,
                    level_data[0], level_data[1], level_data[2], level_data[3],
                    message.channel
                )
                await self.check_and_award_level_roles(
                    message.guild, message.author, level_data[1]
                )

        except Exception as e:
            logger.error(f"Error in on_message XP: {e}")

    @commands.Cog.listener()
    async def on_raw_reaction_add(self, payload: discord.RawReactionActionEvent):
        """Award XP for reactions."""
        if not payload.member or payload.member.bot:
            return

        guild = self.bot.get_guild(payload.guild_id)
        if not guild:
            return

        guild_id = guild.id
        should_notify = False
        level_data = None

        try:
            with db_session_scope() as session:
                db_guild = session.get(Guild, guild_id)
                if not db_guild or not db_guild.xp_enabled:
                    return

                channel = guild.get_channel(payload.channel_id)
                if not XPCog.can_gain_xp(session, guild_id, payload.member, channel):
                    return

                xp_config = XPCog.get_xp_config(session, guild_id)

                db_member = session.get(GuildMember, (guild_id, payload.member.id))
                if not db_member:
                    logger.warning(
                        f"[DUPLICATE TRACKER] xp.on_reaction_add CREATING GuildMember: "
                        f"guild_id={guild_id}, user_id={payload.member.id}, user_id_type={type(payload.member.id)}, "
                        f"display_name={payload.member.display_name}, source=ReactionAddEvent.payload.member"
                    )
                    db_member = GuildMember(
                        guild_id=guild_id,
                        user_id=payload.member.id,
                        display_name=payload.member.display_name,
                    )
                    session.add(db_member)
                    session.flush()

                now = int(time.time())

                # Check if reaction XP tracking is enabled
                if xp_config["track_reactions"] and (now - db_member.last_react_ts) >= xp_config["reaction_cooldown"]:
                    result = XPCog.add_xp(
                        session, guild_id, payload.member.id,
                        xp_config["reaction_xp"], payload.member.display_name, "active",
                        source="reactions"
                    )
                    db_member.last_react_ts = now
                    db_member.reaction_count += 1
                    old_level, new_level, tokens, token_diff = result

                    if new_level > old_level:
                        should_notify = True
                        level_data = (old_level, new_level, tokens, token_diff)

            if should_notify and level_data:
                await self.send_level_up_notification(
                    guild, payload.member,
                    level_data[0], level_data[1], level_data[2], level_data[3],
                    channel  # Reaction channel
                )
                await self.check_and_award_level_roles(guild, payload.member, level_data[1])

        except Exception as e:
            logger.error(f"Error in reaction XP: {e}")

    @commands.Cog.listener()
    async def on_voice_state_update(self, member: discord.Member,
                                     before: discord.VoiceState,
                                     after: discord.VoiceState):
        """Award XP for voice activity."""
        if member.bot:
            return

        guild_id = member.guild.id
        should_notify = False
        level_data = None

        try:
            with db_session_scope() as session:
                db_guild = session.get(Guild, guild_id)
                if not db_guild or not db_guild.xp_enabled:
                    return

                xp_config = XPCog.get_xp_config(session, guild_id)

                db_member = session.get(GuildMember, (guild_id, member.id))
                if not db_member:
                    logger.warning(
                        f"[DUPLICATE TRACKER] xp.on_voice_state_update CREATING GuildMember: "
                        f"guild_id={guild_id}, user_id={member.id}, user_id_type={type(member.id)}, "
                        f"display_name={member.display_name}, source=VoiceState.member"
                    )
                    db_member = GuildMember(
                        guild_id=guild_id,
                        user_id=member.id,
                        display_name=member.display_name,
                    )
                    session.add(db_member)
                    session.flush()

                if db_member.display_name != member.display_name:
                    db_member.display_name = member.display_name

                now = int(time.time())

                # User joined voice
                if before.channel is None and after.channel is not None:
                    if XPCog.can_gain_xp(session, guild_id, member, after.channel):
                        # Award XP for joining voice (with cooldown) - only if voice tracking is enabled
                        if xp_config["track_voice"] and (now - db_member.last_voice_bonus_ts) >= xp_config["voice_interval"]:
                            result = XPCog.add_xp(
                                session, guild_id, member.id,
                                2, member.display_name, "active",
                                source="voice"
                            )
                            db_member.last_voice_bonus_ts = now
                            old_level, new_level, tokens, token_diff = result

                            if new_level > old_level:
                                should_notify = True
                                level_data = (old_level, new_level, tokens, token_diff)

                        db_member.last_voice_join_ts = now
                    else:
                        db_member.last_voice_join_ts = 0

                # User left voice
                elif before.channel is not None and after.channel is None:
                    if db_member.last_voice_join_ts > 0:
                        duration = now - db_member.last_voice_join_ts
                        db_member.voice_minutes += duration // 60

                        # Award XP for time in voice - only if voice tracking is enabled
                        chunks = duration // xp_config["voice_interval"]
                        if xp_config["track_voice"] and chunks > 0:
                            result = XPCog.add_xp(
                                session, guild_id, member.id,
                                xp_config["voice_xp"] * chunks,
                                member.display_name, "active",
                                source="voice"
                            )
                            old_level, new_level, tokens, token_diff = result

                            if new_level > old_level:
                                should_notify = True
                                level_data = (old_level, new_level, tokens, token_diff)

                        db_member.last_voice_join_ts = 0

            if should_notify and level_data:
                await self.send_level_up_notification(
                    member.guild, member,
                    level_data[0], level_data[1], level_data[2], level_data[3],
                    None  # No text channel context for voice XP
                )
                await self.check_and_award_level_roles(member.guild, member, level_data[1])

        except Exception as e:
            logger.error(f"Error in voice XP: {e}")

    @commands.Cog.listener()
    async def on_interaction(self, interaction: discord.Interaction):
        """Award XP for slash command usage."""
        if interaction.type != discord.InteractionType.application_command:
            return
        if interaction.user.bot or not interaction.guild:
            return

        asyncio.create_task(self._award_command_xp(interaction))

    async def _award_command_xp(self, interaction: discord.Interaction):
        """Award XP for slash commands (async)."""
        guild_id = interaction.guild.id
        should_notify = False
        level_data = None

        try:
            with db_session_scope() as session:
                db_guild = session.get(Guild, guild_id)
                if not db_guild or not db_guild.xp_enabled:
                    return

                xp_config = XPCog.get_xp_config(session, guild_id)

                db_member = session.get(GuildMember, (guild_id, interaction.user.id))
                if not db_member:
                    logger.warning(
                        f"[DUPLICATE TRACKER] xp.on_interaction CREATING GuildMember: "
                        f"guild_id={guild_id}, user_id={interaction.user.id}, user_id_type={type(interaction.user.id)}, "
                        f"display_name={interaction.user.display_name}, source=Interaction.user"
                    )
                    db_member = GuildMember(
                        guild_id=guild_id,
                        user_id=interaction.user.id,
                        display_name=interaction.user.display_name,
                    )
                    session.add(db_member)
                    session.flush()

                now = int(time.time())

                if (now - db_member.last_command_ts) >= xp_config["command_cooldown"]:
                    result = XPCog.add_xp(
                        session, guild_id, interaction.user.id,
                        xp_config["command_xp"],
                        interaction.user.display_name, "active"
                    )
                    db_member.last_command_ts = now
                    db_member.command_count += 1
                    old_level, new_level, tokens, token_diff = result

                    if new_level > old_level:
                        should_notify = True
                        level_data = (old_level, new_level, tokens, token_diff)

            if should_notify and level_data:
                member = interaction.guild.get_member(interaction.user.id)
                if member:
                    await self.send_level_up_notification(
                        interaction.guild, member,
                        level_data[0], level_data[1], level_data[2], level_data[3],
                        interaction.channel  # Command channel
                    )
                    await self.check_and_award_level_roles(
                        interaction.guild, member, level_data[1]
                    )

        except Exception as e:
            logger.error(f"Error in command XP: {e}")

    @commands.Cog.listener()
    async def on_presence_update(self, before: discord.Member, after: discord.Member):
        """Award XP for gaming activity (passive)."""
        if after.bot or not after.guild:
            return

        if after.status == discord.Status.offline:
            return

        guild_id = after.guild.id
        should_notify = False
        level_data = None

        try:
            with db_session_scope() as session:
                db_guild = session.get(Guild, guild_id)
                if not db_guild or not db_guild.xp_enabled:
                    return

                xp_config = XPCog.get_xp_config(session, guild_id)

                db_member = session.get(GuildMember, (guild_id, after.id))
                if not db_member:
                    logger.warning(
                        f"[DUPLICATE TRACKER] xp.on_presence_update CREATING GuildMember: "
                        f"guild_id={guild_id}, user_id={after.id}, user_id_type={type(after.id)}, "
                        f"display_name={after.display_name}, source=PresenceUpdate.after"
                    )
                    db_member = GuildMember(
                        guild_id=guild_id,
                        user_id=after.id,
                        display_name=after.display_name,
                    )
                    session.add(db_member)
                    session.flush()

                if db_member.display_name != after.display_name:
                    db_member.display_name = after.display_name

                now = int(time.time())

                # Check if user is playing a game
                is_playing = any(
                    a.type == discord.ActivityType.playing
                    and not getattr(a, "name", "").startswith("Custom Status")
                    for a in after.activities
                )
                was_playing = db_member.last_gaming_ts > 0

                # Started playing
                if is_playing and not was_playing:
                    # Award XP for launching a game - only if game launch tracking is enabled
                    if xp_config["track_game_launch"] and (now - db_member.last_game_launch_ts) >= xp_config["game_launch_cooldown"]:
                        result = XPCog.add_xp(
                            session, guild_id, after.id,
                            2, after.display_name, "passive",
                            source="game_launch"
                        )
                        db_member.last_game_launch_ts = now
                        db_member.last_gaming_ts = now
                        old_level, new_level, tokens, token_diff = result

                        if new_level > old_level:
                            should_notify = True
                            level_data = (old_level, new_level, tokens, token_diff)
                    else:
                        # Even if not awarding XP, still track gaming session start
                        db_member.last_gaming_ts = now

                # Stopped playing
                elif not is_playing and was_playing:
                    duration = now - db_member.last_gaming_ts
                    chunks = duration // xp_config["gaming_interval"]

                    # Award XP for time spent gaming - only if gaming tracking is enabled
                    if xp_config["track_gaming"] and chunks > 0:
                        result = XPCog.add_xp(
                            session, guild_id, after.id,
                            xp_config["gaming_xp"] * chunks,
                            after.display_name, "passive",
                            source="gaming"
                        )
                        old_level, new_level, tokens, token_diff = result

                        if new_level > old_level:
                            should_notify = True
                            level_data = (old_level, new_level, tokens, token_diff)

                    db_member.last_gaming_ts = 0

            if should_notify and level_data:
                await self.send_level_up_notification(
                    after.guild, after,
                    level_data[0], level_data[1], level_data[2], level_data[3],
                    None  # No channel context for gaming/presence XP
                )
                await self.check_and_award_level_roles(after.guild, after, level_data[1])

        except Exception as e:
            logger.error(f"Error in presence XP: {e}")

    # User commands

    @xp.command(name="profile", description="View your XP, level, and tokens")
    @discord.option("member", discord.Member, description="Member to view (default: yourself)", required=False)
    async def xp_profile(
        self,
        ctx: discord.ApplicationContext,
        member: discord.Member = None
    ):
        """View XP profile."""
        target = member or ctx.author

        with db_session_scope() as session:
            db_member = session.get(GuildMember, (ctx.guild.id, target.id))

            if not db_member:
                await ctx.respond(
                    f"{'You have' if target == ctx.author else f'{target.display_name} has'} "
                    "no XP yet! Start chatting to earn XP.",
                    ephemeral=True
                )
                return

            # Get guild settings for token name
            db_guild = session.get(Guild, ctx.guild.id)
            token_name = db_guild.token_name if db_guild and db_guild.token_name else "Hero Tokens"

            embed = discord.Embed(
                title=f"📊 {target.display_name}'s Profile",
                color=discord.Color.gold()
            )

            embed.add_field(name="🏆 Level", value=f"**{db_member.level}**", inline=True)
            embed.add_field(name="⭐ XP", value=f"**{db_member.xp:,.0f}**", inline=True)
            embed.add_field(name=f"🪙 {token_name}", value=f"**{db_member.hero_tokens}**", inline=True)

            embed.add_field(
                name="📈 Activity",
                value=(
                    f"Messages: {db_member.message_count}\n"
                    f"  └ Media: {db_member.media_count}\n"
                    f"Voice: {db_member.voice_minutes} min\n"
                    f"Reactions: {db_member.reaction_count}"
                ),
                inline=True
            )

            embed.set_thumbnail(url=target.display_avatar.url)
            embed.set_footer(text="Earn XP by chatting, voice, reactions, and more!")

        await ctx.respond(embed=embed, ephemeral=True)

    @xp.command(name="leaderboard", description="View server XP leaderboard")
    async def xp_leaderboard(self, ctx: discord.ApplicationContext):
        """View leaderboard."""
        with db_session_scope() as session:
            top_members = (
                session.query(GuildMember)
                .filter(GuildMember.guild_id == ctx.guild.id)
                .order_by(GuildMember.xp.desc())
                .limit(10)
                .all()
            )

            if not top_members:
                await ctx.respond("No one has earned XP yet!", ephemeral=True)
                return

            embed = discord.Embed(
                title=f"🏆 {ctx.guild.name} Leaderboard",
                color=discord.Color.gold()
            )

            leaderboard_text = ""
            medals = ["🥇", "🥈", "🥉"]

            for i, member in enumerate(top_members):
                medal = medals[i] if i < 3 else f"**{i+1}.**"
                name = member.display_name or f"User {member.user_id}"
                leaderboard_text += (
                    f"{medal} {name} - Level **{member.level}** "
                    f"({member.xp:,.0f} XP)\n"
                )

            embed.description = leaderboard_text

        await ctx.respond(embed=embed)

    # Admin commands

    @xp.command(name="give", description="Give XP to a member (Admin)")
    @discord.default_permissions(administrator=True)
    @commands.has_permissions(administrator=True)
    async def xp_give(
        self,
        ctx: discord.ApplicationContext,
        member: discord.Member,
        amount: float
    ):
        """Give XP to a member."""
        if member.bot:
            await ctx.respond("❌ Cannot give XP to bots.", ephemeral=True)
            return

        # PREVENT SELF-GIFTING (Anti-abuse)
        if member.id == ctx.author.id:
            # Log the abuse attempt
            logger.warning(
                f"⚠️ ABUSE ATTEMPT: {ctx.author} (ID: {ctx.author.id}) tried to give themselves "
                f"{amount} XP in guild {ctx.guild.name} (ID: {ctx.guild.id})"
            )
            await ctx.respond(
                "❌ **You cannot give XP to yourself!**\n"
                "This attempt has been logged. Please give XP to other members only.",
                ephemeral=True
            )
            return

        with db_session_scope() as session:
            result = XPCog.add_xp(
                session, ctx.guild.id, member.id,
                amount, member.display_name, "active"
            )
            old_level, new_level, tokens, token_diff = result

        # Log the admin action
        logger.info(
            f"✅ ADMIN GIFT: {ctx.author} gave {amount} XP to {member} "
            f"in guild {ctx.guild.name} (New level: {new_level})"
        )

        await ctx.respond(
            f"✅ Gave **{amount:,.0f} XP** to {member.mention}. "
            f"They are now level **{new_level}**.",
            ephemeral=True
        )

    @xp.command(name="give-tokens", description="Give tokens to a member (Admin)")
    @discord.default_permissions(administrator=True)
    @commands.has_permissions(administrator=True)
    async def xp_give_tokens(
        self,
        ctx: discord.ApplicationContext,
        member: discord.Member,
        amount: int
    ):
        """Give tokens to a member."""
        if member.bot:
            await ctx.respond("❌ Cannot give tokens to bots.", ephemeral=True)
            return

        # PREVENT SELF-GIFTING (Anti-abuse)
        if member.id == ctx.author.id:
            # Log the abuse attempt
            logger.warning(
                f"⚠️ ABUSE ATTEMPT: {ctx.author} (ID: {ctx.author.id}) tried to give themselves "
                f"{amount} tokens in guild {ctx.guild.name} (ID: {ctx.guild.id})"
            )
            await ctx.respond(
                "❌ **You cannot give tokens to yourself!**\n"
                "This attempt has been logged. Please give tokens to other members only.",
                ephemeral=True
            )
            return

        with db_session_scope() as session:
            # Get guild settings for token name
            db_guild = session.get(Guild, ctx.guild.id)
            token_name = db_guild.token_name if db_guild and db_guild.token_name else "Hero Tokens"

            db_member = session.get(GuildMember, (ctx.guild.id, member.id))

            if not db_member:
                logger.warning(
                    f"[DUPLICATE TRACKER] xp.give_tokens CREATING GuildMember: "
                    f"guild_id={ctx.guild.id}, user_id={member.id}, user_id_type={type(member.id)}, "
                    f"display_name={member.display_name}, source=SlashCommand.member"
                )
                db_member = GuildMember(
                    guild_id=ctx.guild.id,
                    user_id=member.id,
                    display_name=member.display_name,
                )
                session.add(db_member)

            db_member.hero_tokens += amount
            final_tokens = db_member.hero_tokens

        # Log the admin action
        logger.info(
            f"✅ ADMIN GIFT: {ctx.author} gave {amount} tokens to {member} "
            f"in guild {ctx.guild.name} (New balance: {final_tokens})"
        )

        await ctx.respond(
            f"✅ Gave **{amount} {token_name}** to {member.mention}. "
            f"They now have **{final_tokens}** tokens.",
            ephemeral=True
        )

    @xp.command(name="export-members", description="Export all Discord members to Excel (Admin)")
    @discord.default_permissions(administrator=True)
    @commands.has_permissions(administrator=True)
    async def xp_export_members(self, ctx: discord.ApplicationContext):
        """Export all guild members to Excel with proper formatting (no scientific notation!)."""
        await ctx.defer(ephemeral=True)

        try:
            # Fetch all members
            members = []
            async for member in ctx.guild.fetch_members(limit=None):
                if not member.bot:
                    members.append({
                        'user_id': str(member.id),
                        'display_name': member.display_name,
                        'username': f'{member.name}',
                        'joined_at': member.joined_at.isoformat() if member.joined_at else ''
                    })

            if not members:
                await ctx.respond("❌ No members found!", ephemeral=True)
                return

            # Create XLSX with TEXT formatted user_id column
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from io import BytesIO
            from datetime import datetime

            wb = Workbook()
            ws = wb.active
            ws.title = "Members"

            # Write headers
            headers = ['user_id', 'display_name', 'username', 'joined_at']
            ws.append(headers)

            # Bold headers
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Write data and format user_id column as TEXT
            for member in members:
                ws.append([
                    member['user_id'],
                    member['display_name'],
                    member['username'],
                    member['joined_at']
                ])

            # Format user_id column as TEXT (prevents Excel scientific notation)
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=1).number_format = '@'  # @ = TEXT format

            # Adjust column widths
            ws.column_dimensions['A'].width = 20  # user_id
            ws.column_dimensions['B'].width = 25  # display_name
            ws.column_dimensions['C'].width = 25  # username
            ws.column_dimensions['D'].width = 25  # joined_at

            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            filename = f"{ctx.guild.name.replace(' ', '_')}_members_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Send as file attachment
            file = discord.File(excel_file, filename=filename)

            await ctx.respond(
                f"✅ Exported **{len(members)}** members to Excel!\n\n"
                f"**Excel-Safe Format:**\n"
                f"• User IDs are pre-formatted as TEXT\n"
                f"• No scientific notation problems!\n"
                f"• Safe to edit and save in Excel\n\n"
                f"Use for backups, analysis, or bulk XP updates via the dashboard.",
                file=file,
                ephemeral=True
            )

            logger.info(f"Exported {len(members)} members to XLSX for guild {ctx.guild.id} by {ctx.author}")

        except Exception as e:
            logger.error(f"Error exporting members: {e}")
            await ctx.respond(f"❌ Error: {str(e)}", ephemeral=True)

    # ═══════════════════════════════════════════════════════════════
    # XP BOOST EVENTS - Background Tasks
    # ═══════════════════════════════════════════════════════════════

    @tasks.loop(minutes=5)
    async def check_expired_boost_events(self):
        """Automatically disable boost events that have expired."""
        from models import XPBoostEvent

        try:
            current_time = int(time.time())

            with db_session_scope() as session:
                # Find active events that have passed their end time
                expired_events = session.query(XPBoostEvent).filter(
                    XPBoostEvent.is_active == True,
                    XPBoostEvent.end_time.isnot(None),
                    XPBoostEvent.end_time < current_time
                ).all()

                for event in expired_events:
                    event.is_active = False
                    logger.info(
                        f"Auto-disabled expired boost event: {event.name} "
                        f"(Guild: {event.guild_id}, ended at {event.end_time})"
                    )

                if expired_events:
                    logger.info(f"Auto-disabled {len(expired_events)} expired boost events")

        except Exception as e:
            logger.error(f"Error checking expired boost events: {e}", exc_info=True)

    @check_expired_boost_events.before_loop
    async def before_check_expired_boost_events(self):
        """Wait for bot to be ready before starting the task."""
        await self.bot.wait_until_ready()

    def cog_unload(self):
        """Clean up when cog is unloaded."""
        self.check_expired_boost_events.cancel()


def setup(bot: commands.Bot):
    bot.add_cog(XPCog(bot))
